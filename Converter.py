from openpyxl import load_workbook
from collections import namedtuple
from fuzzywuzzy import fuzz

wb = load_workbook(filename='./Sample-data-2.xlsx', read_only=True)

#print(str(wb.sheetnames))

ws = wb['Foglio1']

#expected headers, this list is provided by YOU
expectedHeaders = ['clientId', 'reportNo', 'clientContact', 'sampleRef', 'dateReceived', 'sampleId', 'sampleReportNumber', 'clientReference', 'alluminium', 'totalColiforms']

#headers and table data
headers = []
data = []

#grab the data and the headers
for row in ws.rows:
    if len(headers) == 0:
        for cell in row:
            headers.append(cell.value)
    else:
        tmp = []
        for cell in row:
            tmp.append(cell.value)
        data.append(tmp)

#display the data
#print("Data: ")
#print(len(data))

#print("headers")
#for header in headers:
#    print(header)

#print("data")
#for header in headers:
#    print(header)
#    for row in data:
#        print("\t{0}".format(row[headers.index(header)]))

#evaluate the headers
evaluetedHeaders = []
headersPairs = [] #(found, expected)

for expectedHeader in expectedHeaders:
    tmp = []
    for header in headers:
        tmp.append((header, fuzz.ratio(expectedHeader, header)))
    evaluetedHeaders.append(tmp)

for evaluation in evaluetedHeaders:
    #print(evaluation)

    maxElem = evaluation[0]
    for elem in evaluation:
        if maxElem[1] < elem[1]:
            maxElem = elem

    #print(str(maxElem))

    #once the better one has been found delete if from all the other evaluations
    for evaluation2 in evaluetedHeaders:
        for header in evaluation2:
            if header[0] == maxElem[0]:
                evaluetedHeaders[evaluetedHeaders.index(evaluation2)].pop(evaluation2.index(header))

    #print(maxElem[0])

    headersPairs.append((maxElem[0], expectedHeaders[evaluetedHeaders.index(evaluation)]))

#show the headers once evalueted
#print(str(headersPairs))
#print(str(headers))

#pair headers with data
pairedData = []

for row in data:
    tmp = []

    if row:
        for header in headersPairs:
            #print(str(header))
            tmp.append((header[0], row[headers.index(header[0])]))

        pairedData.append(tmp)

#show final paired data
for row in pairedData:
    print(str(row))