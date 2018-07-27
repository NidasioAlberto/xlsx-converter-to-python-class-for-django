from openpyxl import load_workbook
from collections import namedtuple
from fuzzywuzzy import fuzz
import inspect

def takeSecond(elem):
    return elem[1]

def convertXlsxToDjangoClass(GivenClass, filePath, sheetNo=0, headerLimit=-1, expectedHeaders=None, repetitiveColumns=None):
    #0: initializing variables
    allHeaders = []
    headers = []
    data = []
    evaluetedHeaders = [] #[expected][actual]
    sortedEvaluatedHeaders = [] #[expected][actual](index, header score)
    headerPairs = [] #[(expected header index, actual header index), ...]
    pairedData = [] #[(expected header, value), ...]

    #1: get expected headers from the class
    if not expectedHeaders:
        expectedHeaders = [b[0] for b in [a for a in inspect.getmembers(GivenClass, lambda a:not(inspect.isroutine(a))) if not(a[0].startswith('__') and a[0].endswith('__'))]]

    #2: load the worksheet
    wb = load_workbook(filename=filePath, read_only=True)
    ws = wb[wb.sheetnames[sheetNo]]
    
    #3: grab the data and the headers from the worksheet
    for row in ws.rows:
        if len(allHeaders) == 0:
            for cell in row:
                    allHeaders.append(cell.value)
        else:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            if len(tmp) > 0:
                data.append(tmp)
    
    if headerLimit > 0:
        headers = allHeaders[:headerLimit]
    else:
        headers = allHeaders
    
    #4: evaluate the headers
    for expectedHeader in expectedHeaders: #for each expeced header
        tmp = []
        for header in headers: #compare each actual header with the fuzzywuzzy lybrary
            tmp.append(fuzz.ratio(expectedHeader, header))
        evaluetedHeaders.append(tmp)

    #5: organise the cores
    for i in range(0, len(expectedHeaders)):
        scores = [] #(index, header score)
        for j in range(0, len(headers)):
            scores.append((j, evaluetedHeaders[i][j]))    
        scores.sort(key=takeSecond)
        #print(str(scores))
        sortedEvaluatedHeaders.append(scores)

    # for i in range(0, len(headers)):
    #     print("{0}: \t{1}".format(i, headers[i]))

    # print("evaluated headers")
    # for i in range(0, len(evaluetedHeaders)):
    #     print("{0}: \t{1}".format(expectedHeaders[i], str(evaluetedHeaders[i])))

    # print("sorted evaluated headers")
    # for i in range(0, len(sortedEvaluatedHeaders)):
    #     print("{0}: \t{1}".format(expectedHeaders[i], str(sortedEvaluatedHeaders[i])))

    #6: choose the best header for each of the expected
    for i in range(0, len(evaluetedHeaders)):
        for j in range(0, len(headers)):
            k = len(headers) - j - 1#reverse the counter to start from the highest score
            bigger = True
            for l in range(0, len(evaluetedHeaders)):
                #current vs cicle
                if evaluetedHeaders[i][sortedEvaluatedHeaders[i][k][0]] < evaluetedHeaders[l][sortedEvaluatedHeaders[i][k][0]]:
                    bigger = False
            if bigger:
                # print("element found for {0}: {1} {2} score: {3}".format(expectedHeaders[i], headers[sortedEvaluatedHeaders[i][k][0]], sortedEvaluatedHeaders[i][k][0], evaluetedHeaders[i][sortedEvaluatedHeaders[i][k][0]] ))
                # if sortedEvaluatedHeaders[i][k][0] == 140:
                #     print('hey')
                headerPairs.append((i, sortedEvaluatedHeaders[i][k][0])) #(expected header index, actual header index)
                break
            # else:
            #     print("bad position: [{0}][{1}] i:{2} k:{3} j:{4}".format(i, sortedEvaluatedHeaders[i][k][0], i, k, j))

    # print(str(headerPairs))

    #7: pair the expected headers with the data
    for row in data:
        tmp = []
        if row: #check if the row has data in it
            for headerPair in headerPairs:
                tmp1 = expectedHeaders[headerPair[0]]
                tmp3 = headerPair[1]
                #print(tmp3)
                tmp2 = row[tmp3]
                tmp.append((tmp1, tmp2))
            # print(str(tmp))
            pairedData.append(tmp)

    #8: populate the array of the given class with the data
    toReturn = []
    for row in pairedData:
        tmp = GivenClass()
        for elem in row:
            setattr(tmp, elem[0], str(elem[1]))
        toReturn.append(tmp)

    #display the array of classes
    # for elem in toReturn:
    #    print("element:")
    #    for header in expectedHeaders:
    #        print("{0}: {1}".format(header, getattr(elem, header)))

    #9: collect data inside repetitive columns
    if repetitiveColumns != None: #(("Result", "Test ID"), (("chem_value00", "checm_name00"), ...))
        print(str(repetitiveColumns))
        indexes = []
        start = -1
        first = ""
        jumpSize = -1
        for i in range(0, len(allHeaders)):
            # print("current {0} first {1}".format(i, first))
            if first != "":
                if first == allHeaders[i]:
                    jumpSize = i - start
                    break
                if allHeaders[i] in repetitiveColumns[0]:
                    # print("hey {0}".format(allHeaders[i]))
                    indexes.append(repetitiveColumns[0].index(allHeaders[i]))

            if allHeaders[i] in repetitiveColumns[0] and first == "":
                # print("hey {0}".format(allHeaders[i]))
                indexes.append(repetitiveColumns[0].index(allHeaders[i]))
                first = allHeaders[i]
                start = i

        # print(str(indexes))
        # print("start: {0}".format(start))
        # print("first: {0}".format(first))
        # print("jumpSize: {0}".format(jumpSize))
        # print("")

        # for i in range(0, min([len(repetitiveColumns[1])*len(indexes), len(allHeaders) - start])):
        #     print("{0}\t{1}".format(repetitiveColumns[0][indexes[(i) % len(indexes)]], repetitiveColumns[1][int(i/2)][indexes[(i) % len(indexes)]]))

        # print(str(data))
        # print(len(data))
        # print(len(toReturn))

        for j in range(0, len(toReturn)):
            if len(data[j]) >= start + min([len(repetitiveColumns[1])*len(indexes), len(allHeaders) - start]) and len(data[j]) > 0:
                for i in range(0, min([len(repetitiveColumns[1])*len(indexes), len(allHeaders) - start])):
                    #print("\t{0}:\t{1}".format(repetitiveColumns[1][int(i/2)][indexes[(i) % len(indexes)]], data[j][i + start]))

                    #print(j)
                    setattr(toReturn[j], repetitiveColumns[1][int(i/2)][indexes[(i) % len(indexes)]], str(data[j][i + start]))


    return toReturn